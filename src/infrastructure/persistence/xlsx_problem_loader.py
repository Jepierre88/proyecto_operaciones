"""Carga/genera problemas en formato Excel (.xlsx).

Layout del archivo (hoja 'Problema'):

    A1: Nombre del problema     | B1: <texto>
    A2: Tipo de optimizacion    | B2: max | min      (dropdown)
    A3: Coeficiente c1 (de x)   | B3: <numero>
    A4: Coeficiente c2 (de y)   | B4: <numero>
    A5: (vacia)
    A6..E6: encabezados [a1] [a2] [op] [b] [etiqueta]
    A7..E7..: filas de restricciones (la lectura corta en la primera fila vacia)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.application.dto.problem_request import ConstraintRequest, ProblemRequest

SHEET_NAME = "Problema"
HEADER_ROW = 6
FIRST_DATA_ROW = 7
MAX_CONSTRAINT_ROWS = 30  # filas con dropdown de operador pre-aplicado


class XlsxProblemLoader:
    def supports(self, path: Path) -> bool:
        return path.suffix.lower() == ".xlsx"

    # --- lectura ---

    def load(self, path: Path) -> ProblemRequest:
        wb = load_workbook(path, data_only=True)
        sheet = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

        name = str(sheet.cell(row=1, column=2).value or "Problema sin nombre")
        objective_type = str(sheet.cell(row=2, column=2).value or "max").strip().lower()
        c1 = float(sheet.cell(row=3, column=2).value or 0)
        c2 = float(sheet.cell(row=4, column=2).value or 0)

        constraints: list[ConstraintRequest] = []
        row = FIRST_DATA_ROW
        while True:
            a1_cell = sheet.cell(row=row, column=1).value
            a2_cell = sheet.cell(row=row, column=2).value
            op_cell = sheet.cell(row=row, column=3).value
            b_cell = sheet.cell(row=row, column=4).value
            label_cell = sheet.cell(row=row, column=5).value

            if all(v is None or v == "" for v in (a1_cell, a2_cell, op_cell, b_cell)):
                break

            constraints.append(
                ConstraintRequest(
                    a1=float(a1_cell or 0),
                    a2=float(a2_cell or 0),
                    operator=str(op_cell or "<=").strip(),
                    b=float(b_cell or 0),
                    label=str(label_cell).strip() if label_cell else "",
                )
            )
            row += 1

        return ProblemRequest(
            objective_type=objective_type,
            c1=c1,
            c2=c2,
            constraints=constraints,
            name=name,
        )

    # --- generacion de plantilla ---

    def save_template(self, path: Path, with_example: bool = True) -> None:
        wb = Workbook()
        sheet = wb.active
        sheet.title = SHEET_NAME

        self._write_labels(sheet)
        self._write_data_validations(sheet)
        self._write_header_row(sheet)
        self._format_widths_and_borders(sheet)
        if with_example:
            self._write_example_data(sheet)
        self._write_instructions_sheet(wb)

        wb.save(path)

    def _write_labels(self, sheet) -> None:
        bold = Font(bold=True)
        info_fill = PatternFill("solid", fgColor="E3F2FD")

        cells = [
            ("A1", "Nombre del problema"),
            ("A2", "Tipo de optimizacion (max o min)"),
            ("A3", "Coeficiente c1 (de x)"),
            ("A4", "Coeficiente c2 (de y)"),
        ]
        for coord, text in cells:
            sheet[coord] = text
            sheet[coord].font = bold
            sheet[coord].fill = info_fill

    def _write_data_validations(self, sheet) -> None:
        # Dropdown max/min en B2
        dv_type = DataValidation(
            type="list", formula1='"max,min"', allow_blank=False, showDropDown=False
        )
        dv_type.error = 'Debe ser "max" o "min"'
        dv_type.errorTitle = "Valor invalido"
        sheet.add_data_validation(dv_type)
        dv_type.add("B2")

        # Dropdown de operador en columna C, filas 7..(7+MAX-1)
        dv_op = DataValidation(
            type="list", formula1='"<=,>=,="', allow_blank=True, showDropDown=False
        )
        dv_op.error = 'Debe ser "<=", ">=" o "="'
        dv_op.errorTitle = "Operador invalido"
        sheet.add_data_validation(dv_op)
        last_row = FIRST_DATA_ROW + MAX_CONSTRAINT_ROWS - 1
        dv_op.add(f"C{FIRST_DATA_ROW}:C{last_row}")

    def _write_header_row(self, sheet) -> None:
        headers = [
            ("A", "a1 (coef. de x)"),
            ("B", "a2 (coef. de y)"),
            ("C", "Operador"),
            ("D", "b (lado derecho)"),
            ("E", "Etiqueta (opcional)"),
        ]
        thin = Side(border_style="thin", color="888888")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="1565C0")
        header_font = Font(bold=True, color="FFFFFF")
        for col, text in headers:
            cell = sheet[f"{col}{HEADER_ROW}"]
            cell.value = text
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _format_widths_and_borders(self, sheet) -> None:
        widths = {"A": 22, "B": 22, "C": 14, "D": 18, "E": 30}
        for col, w in widths.items():
            sheet.column_dimensions[col].width = w
        sheet.row_dimensions[HEADER_ROW].height = 22

    def _write_example_data(self, sheet) -> None:
        sheet["B1"] = "Cambiar por el nombre del problema"
        sheet["B2"] = "max"
        sheet["B3"] = 3
        sheet["B4"] = 5

        example_rows = [
            (1, 0, "<=", 4, "Planta 1"),
            (0, 2, "<=", 12, "Planta 2"),
            (3, 2, "<=", 18, "Planta 3"),
        ]
        for i, (a1, a2, op, b, lbl) in enumerate(example_rows):
            r = FIRST_DATA_ROW + i
            sheet.cell(row=r, column=1, value=a1)
            sheet.cell(row=r, column=2, value=a2)
            sheet.cell(row=r, column=3, value=op)
            sheet.cell(row=r, column=4, value=b)
            sheet.cell(row=r, column=5, value=lbl)

    def _write_instructions_sheet(self, workbook: Workbook) -> None:
        ws = workbook.create_sheet("Instrucciones")
        ws.column_dimensions["A"].width = 100

        bold = Font(bold=True, size=12)
        title = Font(bold=True, size=14, color="1565C0")

        lines = [
            ("Como usar esta plantilla", title),
            ("", None),
            ("1. En la hoja 'Problema', completa los datos de la cabecera (filas 1 a 4).", None),
            ("   - B1: nombre del problema (texto libre)", None),
            ("   - B2: tipo de optimizacion (elige del menu: max o min)", None),
            ("   - B3: coeficiente c1 (acompana a x en la funcion objetivo)", None),
            ("   - B4: coeficiente c2 (acompana a y en la funcion objetivo)", None),
            ("", None),
            ("2. Desde la fila 7 ingresa cada restriccion (una por fila):", None),
            ("   - Columna A: a1 (coeficiente de x)", None),
            ("   - Columna B: a2 (coeficiente de y)", None),
            ("   - Columna C: operador (<=, >= o =)  --  hay un menu desplegable", None),
            ("   - Columna D: b (termino del lado derecho)", None),
            ("   - Columna E: etiqueta (opcional, aparece en la grafica)", None),
            ("", None),
            ("3. Importante: NO agregues las restricciones x >= 0 ni y >= 0.", None),
            ("   La aplicacion las incluye automaticamente (no-negatividad).", None),
            ("", None),
            ("4. Guarda el archivo como .xlsx y cargalo en la app con:", None),
            ("   Archivo > Cargar problema...", None),
            ("", None),
            ("Ejemplo cargado en la plantilla:", bold),
            ("   Max Z = 3x + 5y", None),
            ("   s.a.  x <= 4", None),
            ("         2y <= 12", None),
            ("         3x + 2y <= 18", None),
            ("         x, y >= 0   (automatico, NO escribir)", None),
            ("   Optimo esperado: (x=2, y=6) con Z = 36", None),
        ]
        for i, (text, font) in enumerate(lines, start=1):
            cell = ws.cell(row=i, column=1, value=text)
            if font is not None:
                cell.font = font
